import re
import glob
from openpyxl import Workbook

from .parser import FormulaParser


def extract_viscosity(formula_path: str) -> list:
    """
    提取粘度数据
    :param formula_path:
    :return:
    """
    all_formula_files = glob.glob(f"{formula_path}/**/*.xlsx", recursive=True)

    result = []
    for filepath in all_formula_files:
        parser = FormulaParser(filepath)
        formulas = parser.parse()
        for formula in formulas:
            after_adding_requirement = formula['metas']['after_adding_requirement']
            if isinstance(after_adding_requirement, list):
                for requirement in after_adding_requirement:
                    if requirement.find("粘度要求") >= 0:
                        result.append(dict(name=formula['name'], viscosity=requirement))
            elif after_adding_requirement and isinstance(after_adding_requirement, str):
                if after_adding_requirement.find("粘度要求") >= 0:
                    result.append(dict(name=formula['name'], viscosity=after_adding_requirement))

    return _extract_value(result)


def _extract_value(data: list) -> list:
    """
    :param data:
    :return: List[dict(viscosity, name, min, max)]
    """
    # .+? 非贪婪模式  .+ 贪婪模式
    pattern1 = re.compile(r'.+?(\d+)\s*±\s*(\d+)\s*d[P|p]a.*')
    pattern2 = re.compile(r'.+?(\d+)\s*~\s*(\d+)\s*d[P|p]a.*')

    result = []
    for item in data:
        viscosity = item['viscosity'].replace('～', '~').replace('-', '~')
        match = pattern1.match(viscosity)
        if match:
            middle, extent = match.groups()
            result.append({
                "name": item["name"],
                "min": int(middle) - int(extent),
                "max": int(middle) + int(extent),
            })
            continue

        match = pattern2.match(viscosity)
        if match:
            _min, _max = match.groups()
            result.append({
                "name": item["name"],
                "min": int(_min),
                "max": int(_max),
            })
            continue

        result.append(item)

    return result


def viscosity2excel(formula_path: str, viscosity_filepath: str):
    data = extract_viscosity(formula_path)

    wb = Workbook()
    ws = wb.active

    for index, item in enumerate(data):
        ws.cell(row=index + 1, column=1, value=item['name'])
        if 'min' in item:
            ws.cell(row=index + 1, column=2, value=item['min'])
            ws.cell(row=index + 1, column=3, value=item['max'])
        if 'viscosity' in item:
            ws.cell(row=index + 1, column=4, value=item['viscosity'])

    wb.save(filename=viscosity_filepath)


def viscosity2db(formula_path: str, host, user, password, dbname):
    data = extract_viscosity(formula_path)

    # todo to database


def extract_fineness():
    """提取细度数据"""
    pass
