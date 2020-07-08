import copy
import logging
import os
import re
from datetime import datetime

from openpyxl import load_workbook

logger = logging.getLogger("work4rd")


def new_formula() -> dict:
    return {
        'name': '',
        'created_at': '',
        'version': '',
        'category': '',
        'common_name': '',
        'description': '',

        'materials': [
            # dict(name='', amount='', unit='kg', workshop='', memo=''),
        ],

        'extend_materials': [
            # dict(name='', amount='', unit='%', workshop='', memo=''),
        ],

        'technologies': {  # 工艺要求
            'grind_times': '',  # 研磨次数
            'grind_temperature': '<=45',  # 出料温度要求
            'grind_machine': '',  # 研磨设备
            'grind_granule': '<=20um',  # 研磨细度
            'grind_speed': '',  # 研磨速度

            'viscosity': '',  # 260~270dpas/25℃

            'package_way': '',  # 包装过滤方式
            'package_screen': '',  # 过滤袋规格 100T
            'package_specification': '',  # 5kg
            'package_ratio': '',  # 3:1
            'package_part_b': '',  # HD21
            'package_label': '',  # 标签要求 160dpas,两头贴
        },

        'metas': {
            'mixing_requirement': '',  # 配料要求
            'grind_requirement': '',  # 研磨要求
            'after_adding_requirement': '',  # 加料要求
            'package_requirement': '',  # 包装其他要求
        },
    }


class FormulaParser(object):
    """ 分析一个excel文件，获取产品配方信息"""

    def __init__(self, filepath: str):
        self.filepath = filepath
        self.formulas = []

        try:
            self.workbook = load_workbook(self.filepath, data_only=True)
            self.worksheets_num = len(self.workbook.worksheets)
        except Exception as e:
            logger.error(f"Loading {self.filepath} Error: {e}")
            self.workbook = None
            self.worksheets_num = None

    def parse(self) -> list:
        """ return all formulas in self.workbook
        formula is a dict @see new_formula()
        """
        if self.workbook is None:
            return self.formulas

        _formula = self.parse_filename()
        materials = self.get_mixing_materials()
        products = self.get_products()

        for product_name, sheet_name in products:
            formula = copy.deepcopy(_formula)  # 这里必须 copy 一份，防止引用
            formula['name'] = product_name  # 重置为车间使用的产品名
            formula['materials'] = materials

            # 加料信息
            extends_info = self.get_extends_info(sheet_name)
            formula['extend_materials'] = extends_info['materials']
            formula['metas']['after_adding_requirement'] = extends_info['requirements']

            # 包装信息
            formula['technologies'].update(self.get_package_info(sheet_name))

            # 研磨信息
            formula['technologies'].update(self.get_grind_info(sheet_name))

            self.formulas.append(formula)

        return self.formulas

    def parse_filename(self) -> dict:
        """
        解析文件名
        return dict formula update dict(name, created_at, description, version)
        """
        filename = os.path.basename(self.filepath).replace("（", "(").replace("）", ")")
        pattern = re.compile(r'''^(?P<created_at>\d{4}-\d{1,2}-\d{1,2})?
                                  (?P<name>.+)
                                  \(
                                  (?P<version>B-\d{1,2})
                                  \)
                                  (?P<description>.*)
                                  \.xlsx$''', re.X)
        match = pattern.match(filename)
        if match:
            formula = match.groupdict()
            formula['name'] = formula['name'].strip().strip("_")
        else:
            formula = dict(name=os.path.splitext(filename)[0],
                           created_at=datetime.utcnow().strftime("%Y-%m-%d"),
                           version='B-01',
                           description='')

        _formula = new_formula()
        _formula.update(formula)

        return _formula

    def get_products(self) -> list:
        """ return a [(product_name, sheet_name)] list """
        if self.workbook is None:
            return []

        products = []
        for sheet in self.workbook.sheetnames:
            if sheet.find('配料单') < 0:  # 跟踪单
                flag = self.workbook[sheet]["A4"].value
                name = self.workbook[sheet]["B4"].value
                if flag != "品名" or not name:
                    logger.warning(f'在 {self.filepath} 的 sheet {sheet} 中没有找到产品名, '
                                   '请检查是否为不标准的配方格式。')
                    continue

                products.append((name, sheet))
        return products

    def get_mixing_materials(self, start_row=8) -> list:
        """
        获取配料表
        return a list[dict(name, amount, workshop)]
        """
        materials = []
        try:
            ws = self.workbook['配料单']
        except KeyError:
            ws = self.workbook.worksheets[0]

        for row in ws[f"B{start_row}:C{ws.max_row}"]:
            name, amount = [cell.value for cell in row]
            if name and (isinstance(amount, float) or isinstance(amount, int)):
                materials.append(dict(name=str(name), amount=amount, unit='kg', workshop='配料'))

        return materials

    def get_extends_info(self, sheet: str) -> dict:
        """
        获取加料信息
        return a dict(materials<list>, requirements<list>)
        """
        info = dict(materials=[], requirements=[])
        ws = self.workbook[sheet]
        title = ws["A2"].value
        if title is None or (isinstance(title, str) and title.replace(" ", "") == "RoHS配料生产记录表"):  # 配料单
            return info

        start_row = self.get_extends_start_row(ws)

        for row in range(start_row, ws.max_row):
            name = ws.cell(row=row, column=1).value
            amount = ws.cell(row=row, column=3).value
            if name == '返回油墨':  # 加料信息结束行
                break

            if name:
                if isinstance(amount, float) or isinstance(amount, int):
                    info['materials'].append(dict(name=name, amount=amount, unit='%', workshop='加料'))
                elif amount == '稀释剂':
                    info['materials'].append(dict(name=name, amount=0, unit='kg', workshop='加料'))
                elif not amount:
                    info['requirements'].append(name)

        return info

    def get_package_info(self, sheet: str) -> dict:
        info = {
            'package_way': '',  # 包装过滤方式
            'package_screen': '',  # 过滤袋规格 100T
            'package_specification': '',  # 5kg
            'package_ratio': '',  # 3:1
            'package_part_b': '',  # HD21
            'package_label': '',  # 标签要求 160dpas,两头贴
        }
        ws = self.workbook[sheet]
        start_row = self.get_package_start_row(ws)
        text = ws[f"I{start_row}"].value

        if not text or not isinstance(text, str):
            logger.warning(f"读取不到包装信息: {self.filepath} {sheet}")
            return info

        # 获取标签信息
        label_pattern = re.compile(r'标签[：:](.+)')
        label_match = label_pattern.match(text)
        if label_match:
            info['package_label'] = label_match.group(1)

        # 获取包装规格信息
        spec_pattern = re.compile(r'\d+\s*kg', re.M | re.I)
        spec_match = spec_pattern.search(text)
        if spec_match:
            info['package_specification'] = spec_match.group()

        # 获取B组分信息
        part_b_pattern = re.compile(r'[HD][DBASHF试样无卤素\d\-（）()]+', re.M | re.I)
        part_b_match = part_b_pattern.search(text)
        if part_b_match:
            info['package_part_b'] = part_b_match.group()

        # 获取配比
        ratio_pattern = re.compile(r'A[/:：]B=(\d+[/:：]\d+)', re.M | re.I)
        ratio_match = ratio_pattern.search(text)
        if ratio_match:
            info['package_ratio'] = ratio_match.group(1)

        # 获取过滤方式
        way_pattern = re.compile(r'过滤方式[:：](.*)', re.M | re.I)
        way_match = way_pattern.search(text)
        if way_match:
            info['package_way'] = way_match.group(1)

        # 获取滤袋规格
        screen_pattern = re.compile(r'滤袋规格[:：](.*)', re.M | re.I)
        screen_match = screen_pattern.search(text)
        if screen_match:
            info['package_screen'] = screen_match.group(1)

        return info

    def get_grind_info(self, sheet_name: str):
        """获取研磨信息"""
        info = {
            'grind_times': '',  # 研磨次数
            'grind_temperature': '<=45',  # 出料温度要求
            'grind_machine': '',  # 研磨设备
            'grind_granule': '<=20um',  # 研磨细度
            'grind_speed': '',  # 研磨速度
        }
        ws = self.workbook[sheet_name]
        text = ws["I8"].value

        if not text or not isinstance(text, str):
            logger.warning(f"读取不到研磨信息: {self.filepath} {sheet_name}")
            return info

        # 研磨次数
        pattern = re.compile(r'(\d|\d[\-~]\d)遍', re.M)
        match = pattern.search(text)
        if match:
            info['grind_times'] = match.group(1)

        # 温度要求
        pattern = re.compile(r'出料温度[:：](.*℃)', re.M)
        match = pattern.search(text)
        if match:
            info['grind_temperature'] = match.group(1)

        # 细度要求
        pattern = re.compile(r'细度[:：](.*[μu]m)', re.M)
        match = pattern.search(text)
        if match:
            info['grind_granule'] = match.group(1)

        # 速度要求
        pattern = re.compile(r'流量[:：](.*kg/h)', re.M)
        match = pattern.search(text)
        if match:
            info['grind_speed'] = match.group(1)

        return info

    @staticmethod
    def get_extends_start_row(ws) -> int:
        """获取加料信息起始行"""
        start_row = 10
        for row in range(start_row, ws.max_row):
            if ws.cell(row=row, column=1).value == "物料名称":
                return row

        return start_row

    @staticmethod
    def get_package_start_row(ws) -> int:
        """获取包装信息起始行"""
        start_row = 20
        for row in range(start_row, ws.max_row):
            if ws.cell(row=row, column=1).value == "包装工序":
                return row

        return start_row
