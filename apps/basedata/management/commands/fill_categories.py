import os

from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from openpyxl import load_workbook

from apps.basedata.models import MaterialCategory


class Command(BaseCommand):
    help = 'fill MaterialCategory 填充类别表基础数据'

    def handle(self, *args, **options):
        filepath = os.path.join(
            os.path.join(settings.BASE_DIR, "storage"),
            "类别.xlsx"
        )

        if not os.path.exists(filepath):
            raise CommandError(f"{filepath} 文件不存在")

        wb = load_workbook(filepath, read_only=True)
        ws = wb["类别"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=4, values_only=True):
            code, spec, name, _ = row
            if not spec:
                spec = ''
            c = MaterialCategory(code=code, spec=spec, name=name)
            c.save()

        self.stdout.write(self.style.SUCCESS("SUCCESS!!!"))
